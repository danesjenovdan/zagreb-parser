import logging
import re
from datetime import datetime, timedelta

import openpyxl
import requests
from pypdf import PdfReader

from .base_parser import BaseParser

logger = logging.getLogger("session logger")


class VotesParser(BaseParser):
    def __init__(self, item, storage):
        """
        {
            "vote_name": ["Davanje prethodne suglasnosti na Prijedlog odluke o izmjenama Statuta Centra za kulturu i film Augusta Cesarca"],
            "champions": ["Upravno vijeće Centra za kulturu i film Augusta Cesarca", "Odbor za kulturu, međugradsku i međunarodnu suradnju i civilno društvo", "Odbor za Statut, Poslovnik i propise"],
            "links": [
            {
                "href": "https://web.zagreb.hr/Sjednice/2021/sjednice_skupstine_2021.nsf/0/7AD39F407F4407C4C1258C20002D0760/$FILE/00 Dopis ureda.pdf",
                "text": "DOPIS GRADSKOG UREDA"
            }, {
                "href": "https://web.zagreb.hr/Sjednice/2021/sjednice_skupstine_2021.nsf/0/54BE9F16DABE437EC1258C20002D0CD8/$FILE/01 Dopis predlagatelja.pdf",
                "text": "DOPIS PREDLAGATELJA"
            }, {
                "href": "https://web.zagreb.hr/Sjednice/2021/sjednice_skupstine_2021.nsf/0/6AD1216C4D6FFCD7C1258C20002D1210/$FILE/Prijedlog odluke o izmjenama Statuta Statuta A. Cesarec 2025.pdf",
                "text": "PRIJEDLOG ODLUKE: o izmjenama Statuta"
            }, {
                "href": "https://web.zagreb.hr/Sjednice/2021/sjednice_skupstine_2021.nsf/0/4771D0C19FB0DB25C1258C20002D5511/$FILE/STATUT-CKF A. Cesarec 2022.pdf",
                "text": "STATUT"
            }, {
                "href": "https://web.zagreb.hr/Sjednice/2021/sjednice_skupstine_2021.nsf/0/A2F5F8F0152D4E3FC1258C20002D5D32/$FILE/Odluka-o-izmjenama-i-dopunama-Statuta-Centra_2024.pdf",
                "text": "ODLUKA: o izmjenama i dopunama 2024."
            }],
            "session_text": "41. sjednica Gradske skupštine Grada Zagreba, utorak 18.02.2025 u 09:00h, u Staroj gradskoj vijećnici, Ulica sv. Ćirila i Metoda 5/I., dvorana \"A\".",
            "no_agenda": "TOČKA: 10.",
            "url": "https://web.zagreb.hr/Sjednice/2021/sjednice_skupstine_2021.nsf/PW_NEW?OpenForm&ParentUNID=D2971AF9A26ED6DDC1258C20002CD136&font=14"
        }
        """
        # call init of parent object
        super(VotesParser, self).__init__(storage)
        logger.info(".:Vote PARSER:.")
        logger.debug(item)

        self.log_file = open("parser_log.txt", "a")

        organization = self.storage.organization_storage.get_or_add_object(
            {
                "name": self.parse_organization(item.get("session_text")),
            }
        )

        self.session = storage.session_storage.get_or_add_object(
            {
                "name": self.parse_session_name(item.get("session_text")),
                "organization": organization.id,
                "organizations": [organization.id],
                "in_review": False,
                "classification": "regular",
                "mandate": storage.mandate_id,
            }
        )

        session_start_time = self.parse_start_time(item.get("session_text"))
        if self.session.start_time is None and session_start_time:
            self.session.update_start_time(session_start_time)
        session_no = self.get_no_session(item.get("session_text"))
        logger.info(self.parse_session_name(item.get("session_text")))
        file_name = f"files/{session_no}. sjednica GSGZ.xlsx"
        logger.info(file_name)
        try:
            xml_votes = BallotsXLSParser().parse_file(file_name)
        except FileNotFoundError:
            print("FileNotFoundError")
            return

        self.log_file.write(
            f"\n _________________________________________________________________________________ \n"
        )

        self.log_file.write(
            f"\n ----> {self.parse_session_name(item.get('session_text'))} <---- \n"
        )
        url = f"https://web.zagreb.hr/sjednice/2021/sjednice_skupstine_2021.nsf/DRJ?OpenAgent&{session_no}"
        self.log_file.write(f"\nLink do sjednice: {url} \n")

        item["votes"].sort(key=lambda x: x.get("order"))

        for vote_item in item["votes"]:
            vote_name = " ".join(vote_item.get("vote_name"))

            if self.session.vote_storage.check_if_motion_is_parsed(
                {"text": vote_name, "datetime": session_start_time.isoformat()}
            ):
                print("Skip motion")
                continue

            order = int(vote_item.get("order"))

            enums = re.search(r"((\d+)\))?\s(.+)", vote_item.get("url_text")).groups()
            sub_agenda = enums[1]
            xml_vote_matches = []

            for idx, temp_xml_vote in enumerate(xml_votes):
                print(vote_item.get("no_agenda"), sub_agenda)
                if temp_xml_vote["ai_order"] == vote_item.get("no_agenda"):
                    if sub_agenda:
                        if temp_xml_vote["ai_sub_order"] == sub_agenda:
                            xml_vote_matches.append(temp_xml_vote)
                            xml_votes[idx]["parsed"] = True
                    else:
                        xml_vote_matches.append(temp_xml_vote)
                        xml_votes[idx]["parsed"] = True

            agenda_name = self.agenda_item_name(
                vote_name, vote_item.get("no_agenda"), sub_agenda
            )
            start_time = (
                session_start_time + timedelta(minutes=int(order))
            ).isoformat()
            ai = self.save_agenda_item(order, agenda_name, start_time)

            if not ai.is_new:
                print("Skip agenda item")
                continue

            if not xml_vote_matches:

                ai_parladata_url = f"https://parladata-zagreb.lb.djnd.si/admin/parladata/agendaitem/{ai.id}/change/"
                print("NO MATCH")
                print("...:::::!!!!!!:::::.....")
                print(vote_name)
                # self.log_file.write(f"Session {session_no}, vote: {vote_name} not found in xls by agenda number\n")

                self.log_file.write(
                    f"""
Dodana je točka dnevnega reda brez glasovanja: {agenda_name}
Povezava na točko: {ai_parladata_url}
"""
                )

                for link in vote_item.get("links"):
                    self.storage.parladata_api.links.set(
                        {
                            "agenda_item": ai.id,
                            "url": link.get("href").replace(" ", "+"),
                            "name": link.get("text"),
                            "note": link.get("text"),
                        }
                    )

                print("...:::::!!!!!!:::::.....")
                continue

            if len(xml_vote_matches) == 1:
                self.save_vote(
                    xml_vote_matches[0],
                    vote_item,
                    vote_name,
                    self.session,
                    start_time,
                    ai,
                )
            else:
                for xml_vote in xml_vote_matches:
                    vote_name = xml_vote["name"]
                    self.save_vote(
                        xml_vote, vote_item, vote_name, self.session, start_time, ai
                    )

        for xml_vote in xml_votes:
            if not xml_vote.get("parsed"):
                name = xml_vote.get("name", f"Glasanje bez imena: {xml_vote['order']}")
                motion_obj = self.save_vote(
                    xml_vote, None, name, self.session, start_time, None
                )
                motion_url = f"https://parladata-zagreb.lb.djnd.si/admin/parladata/motion/{motion_obj.id}/change/"
                self.log_file.write(
                    f"""
Nepovezano glasovanje: {xml_vote["name"]}
Glasanje je bilo {xml_vote["order"]}. u datoteki
Povezava na motion: {motion_url}
                    """
                )

        self.log_file.close()

    def save_agenda_item(self, order, name, start_time):
        ai = self.session.agenda_items_storage.get_or_add_object(
            {
                "datetime": start_time,
                "name": name,
                "session": self.session.id,
                "order": order + 1,
                # "gov_id": vote_item.get ("url")
            }
        )
        return ai

    def save_vote(self, xml_vote, vote_item, vote_name, session, start_time, ai):
        motion_data = {
            "session": self.session.id,
            "text": vote_name,
            "title": vote_name,
            "datetime": start_time,
            "agenda_items": [ai.id] if ai else [],
            "epa": None,
            "law": None,
            "result": self.parse_results(xml_vote["result"]),
        }
        motion_obj = self.session.vote_storage.get_or_add_object(motion_data)
        self.save_ballots(xml_vote, motion_obj)
        if vote_item:
            for link in vote_item.get("links"):
                self.storage.parladata_api.links.set(
                    {
                        "agenda_item": ai.id,
                        "motion": motion_obj.id,
                        "url": link.get("href").replace(" ", "+"),
                        "name": link.get("text"),
                        "note": link.get("text"),
                    }
                )
        return motion_obj

    def agenda_item_name(self, name, no_agenda, sub_agenda):
        if sub_agenda:
            sub_agenda = f"{sub_agenda.lstrip("0")}."
        else:
            sub_agenda = ""
        return f"Točka {no_agenda}.{sub_agenda} {name}"

    def parse_results(self, text):
        if text == "Prijedlog usvojen":
            return True
        else:
            return False

    def save_ballots(self, xml_vote, motion):
        if motion.is_new:
            ballots = xml_vote["ballots"]
            data = []
            for ballot in ballots:
                if ballot["name"] in ["Predsjedavajući", ".  ()"]:
                    continue
                voter = self.storage.people_storage.get_or_add_object(
                    {"name": ballot.pop("name")}
                )
                ballot["vote"] = motion.vote.id
                ballot["personvoter"] = voter.id
                data.append(ballot)
            self.session.vote_storage.set_ballots(data)

    def get_no_session(self, text):
        return re.search(r"(\d+).", text).group(1)

    def parse_session_name(self, text):
        return re.search(r"([0-9]{1,2})(\. sjednica) ([a-zA-Zš ]*)", text).group(0)

    def parse_organization(self, text):
        return re.search(r"([0-9]{1,2})(\. sjednica) ([a-zA-Zš ]*)", text).group(3)

    def parse_start_time(self, text):
        groups = re.search(
            r"([0-9]{2}\.[0-9]{2}\.[0-9]{4})\su\s([0-9]{2}:[0-9]{2})h", text
        )
        return datetime.strptime(groups[0], "%d.%m.%Y u %H:%Mh")

    def get_order(self, text):
        return re.search(r"\d+", text).group(0)


class BallotsXLSParser:
    def parse_file(self, file_path="files/1. sjednica GSGZ.xlsx"):
        wb_obj = openpyxl.load_workbook(file_path)
        sheet_obj = wb_obj.active
        self.file_path = file_path

        max_col = sheet_obj.max_column
        m_row = sheet_obj.max_row

        data = []

        for i in range(1, m_row + 1):
            row = [sheet_obj.cell(row=i, column=j).value for j in range(1, max_col + 1)]
            data.append(row)

        return self.parse_page(data)

    def parse_page(self, page_lines):
        session_name_regex = r"(\d+|\*)\s*((\d+)\))?\s(.+)"
        output = []
        print("parsing page")
        single_vote = {"ballots": []}
        self.state = "META"
        order = 1
        for row in page_lines:
            print(self.state, row)
            if (
                not any(row)
                or row[0] == "Session Name"
                or row[0] in ["Naziv točke dnevnog reda", "Naziv tocke dnevnog reda"]
                and single_vote["ballots"]
            ):
                # next vote
                if not "name" in single_vote:
                    continue
                single_vote["file"] = self.file_path
                # print(single_vote)
                try:
                    re_name = re.search(
                        session_name_regex, single_vote["name"]
                    ).groups()
                except AttributeError:
                    print("ERROR no group")
                    # print(single_vote)
                    re_name = ["", "", "", single_vote["name"]]

                single_vote["name"] = re_name[3]
                single_vote["ai_order"] = re_name[0]
                single_vote["ai_sub_order"] = re_name[2]
                single_vote["order"] = order
                order += 1
                # print(single_vote)
                output.append(single_vote)
                # x = input()
                single_vote = {"ballots": []}
                self.state = "META"
            if self.state == "META":
                if row[0] in ["Naziv točke dnevnog reda", "Naziv tocke dnevnog reda"]:
                    self.state = "NAME"
                    continue

            if self.state == "STATS":
                if row[0].strip().startswith("Prijedlog"):
                    single_vote["result"] = row[0].strip().split("\n")[0]
                if row[0].startswith("Ukupni rezultati po strankama"):
                    results = row[0].split("\n")
                    if len(results) > 1:
                        single_vote["result"] = row[0].split("\n")[1]
                    else:
                        continue
                elif (
                    row[0].lower() == "ime" and row[1] == None
                ):  # and row[2].lower() == "za":
                    self.option_indexes = self.get_options_order(row)
                    self.state = "BALLOTS"

            elif self.state == "BALLOTS":
                if (
                    row[0].lower() == "ime" and row[1] == None
                ):  # and row[2].lower() == "za":
                    self.option_indexes = self.get_options_order(row)
                elif row[0].startswith("*"):
                    continue
                else:
                    name = row[0].split("(")[0].strip()
                    single_vote["ballots"].append(
                        {
                            "name": name,
                            "option": self.get_option(self.parse_option(row)),
                        }
                    )
            elif self.state == "NAME":
                if row[0] == "Ukupni rezultat glasovanja":
                    self.state = "STATS"
                else:
                    if "Ukupni rezultat glasovanja" in row[0].split("\n"):
                        # print("merged lines")
                        self.state = "STATS"
                        single_vote["name"] = row[0]
                    else:
                        # print("OK ime")
                        single_vote["name"] = row[0]
        if single_vote["ballots"]:
            single_vote["file"] = self.file_path
            # print(single_vote)
            try:
                re_name = re.search(session_name_regex, single_vote["name"]).groups()
            except AttributeError:
                print("ERROR no group")
                # print(single_vote)
                re_name = ["", "", "", single_vote["name"]]
            single_vote["name"] = re_name[3]
            single_vote["ai_order"] = re_name[0]
            single_vote["ai_sub_order"] = re_name[2]
            single_vote["order"] = order
            output.append(single_vote)
        return output

    def parse_option(self, line):
        try:
            position = line.index("X")
        except:
            return "error"
        for option, index in self.option_indexes.items():
            if position == index:
                return option

    def get_options_order(self, row):
        indexes = {"za": None, "protiv": None, "suzdržan": None, "odsutan": None}
        for key in indexes.keys():
            for i, cell in enumerate(row):
                if cell and cell.lower().startswith(key):
                    indexes[key] = i
        print(indexes)
        return indexes

    def get_option(self, option):
        options_map = {
            "za": "for",
            "protiv": "against",
            "suzdržan": "abstain",
            "odsutan": "absent",
            "error": "absent",
        }
        return options_map[option]
